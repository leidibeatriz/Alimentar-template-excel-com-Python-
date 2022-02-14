from openpyxl import workbook, load_workbook


#Criando a classe pessoa para preenchermos com dados
class Pessoa:
    def __init__(self, nome, profissao, salariobruto,desconto):
        self.nome = nome
        self.professao = profissao
        self.salariobruto = salariobruto
        self.desconto=desconto

    def calcula_salario(self): #função para calcular o salário liquído
        return self.salariobruto - self.desconto


dados = Pessoa('Clara Maria','Veterinária',1500,420) #variável dados recebe valores para os atributos da classe Pessoa

salario_liquido = dados.calcula_salario() #variável salario+liquido recebe o resultado retornado pela função calcula_salario

#abrindo arquivo excel

template_excel = 'preenche.xlsx' #fazendo a variável receber o nome da planilha
wb = load_workbook(template_excel) #carregando a planilha na variável wb

ws = wb.active #aqui pegamos a seção ativa da planilha, caso tenha mais de uma seção utilize wb['NomeSeção']

#inserindo dados na planilha da variavél dados
ws['A3'] = dados.nome
ws['B3'] = dados.professao
ws['C3'] = dados.salariobruto
ws['D3'] = dados.desconto

#salvando os dados na planilha
wb.save(template_excel)

#pegando valores de salário bruto e desconto ca Coluna C e D da Linha 3

sb = ws['C3'].value #esse pega salário bruto
desc = ws['D3'].value #esse pega desconto

ws['E3'] = sb - desc #aqui é calculado e inserido na planilha Coluna E Linha 3

#salvando os dados na planilha novamente
wb.save(template_excel)




